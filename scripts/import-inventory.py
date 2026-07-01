#!/usr/bin/env python3
"""Import inventory Excel file into Supabase with category mapping and stock init."""
import openpyxl
import os, sys, uuid, hashlib
from supabase import create_client, Client

env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            os.environ[k.strip()] = v.strip().strip('"').strip("'")

SUPABASE_URL = os.environ.get("VITE_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("VITE_SUPABASE_ANON_KEY") or os.environ.get("SUPABASE_SERVICE_KEY")
if not SUPABASE_URL or not SUPABASE_KEY:
    print("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY")
    sys.exit(1)

sup: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Map Arabic sheet names to existing categories in the DB
SHEET_CATEGORY_MAP = {
    "قرطاسية": "Stationery",
    "ادوات النظافة": "Cleaning",
    "المطبخ": "Kitchen",
    "ادوات كهربائية": "Electronics",
    "مستلزمات الرياضة": "Sports",
    "طلاء": "Paint",
    "سباكة": "Plumbing",
    "طعام": "Food",
    "اسكراب": "Scrub",
    "اخري": "Other",
}

UNIT_COLS = {"pcs", "box", "doz", "cartoon", "ream", "kg", "set", "pairs",
             "jerrycan", "can (tin)", "bucket", "50kg", "25kg", "20l",
             "10kg", "5kg", "1kg", "500g", "3l", "10g", "5l", "l"}

# Pre-fetch existing categories: name -> id
cat_map = {}
for row in sup.table("inventory_categories").select("id, name").execute().data:
    cat_map[row["name"]] = row["id"]

# Find the latest Excel file
import glob
files = sorted(glob.glob("/home/iico/Downloads/جرد المخزن*.xlsx"), reverse=True)
if not files:
    print("No جرد المخزن Excel file found in Downloads")
    sys.exit(1)
filepath = files[0]
print(f"Reading: {filepath}")

wb = openpyxl.load_workbook(filepath, data_only=True)
total = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    category_name = SHEET_CATEGORY_MAP.get(sheet_name, sheet_name)
    category_id = cat_map.get(category_name)

    if not category_id:
        # Create the category if it doesn't exist
        result = sup.table("inventory_categories").insert({"name": category_name}).execute()
        if result.data:
            category_id = result.data[0]["id"]
            cat_map[category_name] = category_id
        else:
            print(f"  Could not create category '{category_name}', skipping sheet")
            continue

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_name = row[1]
        if not item_name or str(item_name).strip() == "":
            continue
        item_name = str(item_name).strip()

        quantity = 0
        unit = "pcs"
        for i, header in enumerate(headers):
            if i < len(row) and row[i] is not None and header in UNIT_COLS:
                qty = row[i]
                if isinstance(qty, (int, float)) and qty > 0:
                    quantity = int(qty)
                    unit = header
                    break

        if quantity == 0:
            for i in range(2, len(row)):
                if i < len(row) and row[i] is not None and isinstance(row[i], (int, float)) and row[i] > 0:
                    quantity = int(row[i])
                    unit = headers[i] if i < len(headers) and headers[i] in UNIT_COLS else "pcs"
                    break

        if quantity > 0:
            # Generate a deterministic SKU for each item
            raw = f"{category_name}-{item_name}".encode()
            sku = "ITM-" + hashlib.md5(raw).hexdigest()[:8].upper()
            items.append({
                "name": item_name,
                "category_id": category_id,
                "unit": unit,
                "min_stock_level": max(1, quantity // 10),
                "sku": sku,
            })

    if items:
        print(f"{sheet_name} ({category_name}): {len(items)} items")
        for i in range(0, len(items), 100):
            batch = items[i:i+100]
            try:
                # Insert items (skip duplicates on sku)
                result = sup.table("inventory_items").insert(batch, ignore_duplicates="sku").execute()
                new_ids = [item["id"] for item in (result.data or []) if item.get("id")]
                # Initialize stock at 0 for new items
                for item_id in new_ids:
                    sup.table("inventory_stock").insert({"item_id": item_id, "quantity": 0}).execute()
                total += len(new_ids)
                print(f"  Imported {len(new_ids)} items")
            except Exception as e:
                # Fallback: try inserting one by one to handle duplicates
                inserted = 0
                for item in batch:
                    try:
                        result = sup.table("inventory_items").insert(item).execute()
                        if result.data:
                            sup.table("inventory_stock").insert({"item_id": result.data[0]["id"], "quantity": 0}).execute()
                            inserted += 1
                    except Exception:
                        pass
                total += inserted
                print(f"  Imported {inserted}/{len(batch)} items (batch fallback)")

print(f"\nDone: {total} items imported")
